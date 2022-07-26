import urllib
from bs4 import BeautifulSoup
import pandas as pd
import requests
import re
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#取得今年最低本益比
#本益比(英文：Price-to-Earning Ratio 簡稱PE)是價值投資中一定要了解的一種估價(Valuation)方法， 可以幫助你快速判斷股價現在是昂貴或是便宜。
def get_stock_basic_info(stock):
    url = "https://jsjustweb.jihsun.com.tw/z/zc/zca/zca.djhtm?a=" + str(stock)
    html = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(html, 'html.parser')
    tr = soup.find_all('tr')
    per_list = []
    price = []

    for trr in tr:
        tdlist = trr.find_all('td')
        #print("Debug basic info: ", str(tdlist[0].text))
        if str(tdlist[0].text) == "最低本益比" or \
           str(tdlist[0].text) == "年度" :
            new = []
            for i in range(1, len(tdlist)):
                new.append(tdlist[i].text)
            per_list.append(new)
        elif str(tdlist[0].text) == "開盤價":
            #開盤價
            price.append(float(tdlist[1].text))
            #最高價
            price.append(float(tdlist[3].text))
            #最低價
            price.append(float(tdlist[5].text))

    #print("最低本益比(年度):")
    #for i in range(len(per_list[0])):
    #    print("| ", per_list[0][i], " | ", per_list[1][i], " |")

    #公司名稱
    names = soup.find_all('td', class_='t10')

    return (names[0].text.split()[0].replace("()", ""), price, float(per_list[1][0]))

#計算近4個月的總ROE(A)-稅後
#ROE(Return On Equity)
#股東權益報酬率（ Return On Equity，ROE ），是衡量股東權益投資報酬的指標，反映公司利用資產淨值產生獲利的能力。
def get_stock_4_mon_roe(stock):
    url = "https://jsjustweb.jihsun.com.tw/z/zc/zcr/zcr0.djhtm?b=Q&a=" + str(stock)
    roe_list = []
    html = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(html, 'html.parser')
    spans = soup.find_all("div", {"class": "table-row"})
    roe = 0

    for span in spans:
        sspan = span.find_all('span')
        if str(sspan[0].text) == "ROE(A)─稅後" or \
           str(sspan[0].text) == "期別" :
            new = []
            for i in range(1, len(sspan)):
                new.append(sspan[i].text)
            roe_list.append(new)

    for i in range(4):
        roe += float(roe_list[1][i])

    #Show roe for 4 month
    #print("預估ROE: ", roe);
    #print("| 期別 | ROE(A)-稅後 |" )
    #for i in range(4):
    #    print("| ", roe_list[0][i], " | ", roe_list[1][i], " |")

    return float(roe)

#每股淨值(PBR) = 公司的總價值 / 股票數
#取得每股淨值
def get_stock_pbr(stock):
    url = "https://jsjustweb.jihsun.com.tw/z/zc/zcr/zcr0.djhtm?b=Q&a=" + str(stock)
    pbr_list = []
    html = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(html, 'html.parser')
    spans = soup.find_all("div", {"class": "table-row"})
    pbr = 0

    for span in spans:
        sspan = span.find_all('span')
        if str(sspan[0].text) == "期別" and len(pbr_list) == 0:
            new = []
            for i in range(1, len(sspan)):
                new.append(sspan[i].text)
            pbr_list.append(new)
        elif str(sspan[0].text) == "每股淨值(F)(TSE公告數)":
            new = []
            for i in range(1, len(sspan)):
                new.append(sspan[i].text)
            pbr_list.append(new)
            break;

    #Show roe for 4 month
    #print(pbr_list)
    #print("| 期別 | 每股淨值(F)(TSE公告數) |" )
    #for i in range(len(pbr_list[0])):
    #    print("| ", pbr_list[0][i], " | ", pbr_list[1][i], " |")

    return float(pbr_list[1][0])

def main(outfile, stock_list):
    if not stock_list or not outfile:
        print('參數錯誤：股票號碼為空 或 輸出檔名為空')
        return

    #wb = Workbook()
    ws = {}
    #ws.append(['股票名稱', '股票號碼', '股票合理價', '開盤價', '最高價', '最低價', '參考買賣'])

    print("股票名稱,股票號碼,股票合理價, 開盤價, 最高價, 最低價, 參考買賣")

    for i in range(len(stock_list)):
        try:
            stock_company_name, stock_price, stock_year_low_per = get_stock_basic_info(stock_list[i])
            stock_4_mon_roe = get_stock_4_mon_roe(stock_list[i])
            stock_pbr = get_stock_pbr(stock_list[i])
            stock_predict_price = round((stock_year_low_per * stock_4_mon_roe * stock_pbr) / 100, 2)
            if stock_predict_price >= stock_price[2] :
                stock_buy_string = "買"
            else :
                stock_buy_string = " "
            print("%s, %d, %.2f, %.2f, %.2f, %.2f, %s" \
                %(stock_company_name, stock_list[i], stock_predict_price, \
                    stock_price[0], stock_price[1], stock_price[2], \
                    stock_buy_string))
            ws[stock_list[i]] = stock_predict_price
        except:
            print('error: ',stock_company_name)
            continue

    file = open(output_file_name, "w")
    json.dump(ws, file)
    file.close()

if __name__ == '__main__':
    # -------- configurable parameter -------- #
    #bug_list = [2353, ]
    stock_list = \
        [1215, 1605, 1795, \
         2002, 2301, 2303, 2308, 2313, 2317, 2324, 2327, 2330, 2340, 2344, \
         2345, 2357, 2368, 2379, 2382, 2388, 2395, \
         2409, 2412, 2453, 2454, 2458, 2481, 2497, 2603, 2606, 2618, \
         3006, 3008, 3034, 3035, 3037, 3141, 3169, 3189, 3227, 3228, 3406, \
         3443, 3532, 3558, 3592, 3596, 3675, 3680, 3704, 3707, \
         4162, 4755, 4768, 4906, 4919, 4938, 4968, 5269, 5351, 5388, 5425, \
         5471, 5483, 6104, 6138, 6182, 6187, 6217, 6231, 6469, 6488, 6515, \
         6531, 6533, 6719, 6756, 8028, 8046, 8069, 8299, 8478, 9945]
    #stock_list = [1215, 1605]
    output_file_name = 'stock_roe.json'
    # ---------------------------------------- #
    main(output_file_name, stock_list)
